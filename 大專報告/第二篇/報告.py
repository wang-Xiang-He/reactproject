from openpyxl import load_workbook
import openpyxl as opx
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles import Side, Border, colors,Font

# 檢查檔案是否存在
if os.path.isfile(filename):
    print("檔案存在,新增資料")
    wb = load_workbook(filename)
    sheet = wb["數據表"]     
for i in dict_all:
    kwp=dict_all[i]['kwp']
    name=dict_all[i]['name']
    namelis.append(name)
    kwplis.append(kwp)
titles=[]
titles.append('資料日期')
for i in range(len(kwplis)):
    titles.append(namelis[i])
    titles.append("總建置容量"+str(kwplis[i])+' (kWp)')
    titles.append(" ")
    titles.append(" ")
sheet.append(titles)
titles=[]
titles.append('年-月-日')
for i in range(len(dict_all)):
    titles.append('當日發電量(kWh)')
    titles.append('容因%')
    titles.append('PR值')
    titles.append('模組溫度')
sheet.append(titles)
             

auto(wb,sheet,filename)

